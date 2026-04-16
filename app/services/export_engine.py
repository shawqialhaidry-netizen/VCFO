"""
export_engine.py — Phase 18 (Localized)
Builds exportable artifacts (Excel, JSON bundle) from analysis data.
Fully localized via embedded translation table.
"""
from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── Embedded translation table (subset of app i18n) ──────────────────────────
_TR: dict[str, dict[str, str]] = {
  "revenue": {
    "en": "Revenue",
    "ar": "الإيرادات",
    "tr": "Gelir"
  },
  "cogs": {
    "en": "Cost of Goods Sold",
    "ar": "تكلفة البضاعة المباعة",
    "tr": "Satılan Malın Maliyeti"
  },
  "gross_profit": {
    "en": "Gross Profit",
    "ar": "الربح الإجمالي",
    "tr": "Brüt Kâr"
  },
  "operating_profit": {
    "en": "Operating Profit",
    "ar": "ربح التشغيل",
    "tr": "Faaliyet Kârı"
  },
  "net_profit": {
    "en": "Net Profit",
    "ar": "صافي الربح",
    "tr": "Net Kar"
  },
  "exp_section_title": {
    "en": "Expense Breakdown",
    "ar": "تحليل المصروفات",
    "tr": "Gider Analizi"
  },
  "exp_section_subtitle": {
    "en": "What is driving your costs",
    "ar": "ما الذي يحرّك تكاليفك",
    "tr": "Maliyetlerinizi ne yönlendiriyor"
  },
  "exp_cogs_title": {
    "en": "Cost of Goods Sold",
    "ar": "تكلفة البضاعة المباعة",
    "tr": "Satılan Mal Maliyeti"
  },
  "exp_opex_title": {
    "en": "Operating Expenses",
    "ar": "مصروفات التشغيل",
    "tr": "Faaliyet Giderleri"
  },
  "exp_combined_title": {
    "en": "All Cost Items",
    "ar": "جميع بنود التكاليف",
    "tr": "Tüm Maliyet Kalemleri"
  },
  "exp_pct_revenue": {
    "en": "% of Revenue",
    "ar": "% من الإيرادات",
    "tr": "Gelirin %'si"
  },
  "exp_pct_costs": {
    "en": "% of Total Costs",
    "ar": "% من إجمالي التكاليف",
    "tr": "Toplam Maliyetin %'si"
  },
  "exp_trend": {
    "en": "MoM",
    "ar": "شهري",
    "tr": "Aylık"
  },
  "exp_top_item": {
    "en": "Largest cost item",
    "ar": "أكبر بند تكلفة",
    "tr": "En büyük maliyet kalemi"
  },
  "exp_total_costs": {
    "en": "Total Cost Load",
    "ar": "إجمالي التكاليف",
    "tr": "Toplam Maliyet Yükü"
  },
  "exp_item_count": {
    "en": "{n} items",
    "ar": "{n} بنود",
    "tr": "{n} kalem"
  },
  "exp_flag_cogs_high": {
    "en": "COGS at {pct}% of revenue — above healthy threshold.",
    "ar": "تكلفة البضاعة {pct}% من الإيرادات — فوق الحد المقبول.",
    "tr": "SMM gelirin {pct}%'i — sağlıklı eşiğin üzerinde."
  },
  "exp_flag_item_surge": {
    "en": "{name} rose {mom}% this month ({pct}% of revenue).",
    "ar": "{name} ارتفع {mom}% هذا الشهر ({pct}% من الإيرادات).",
    "tr": "{name} bu ay {mom}% arttı (gelirin {pct}%'i)."
  },
  "exp_flag_concentration": {
    "en": "{name} represents {pct}% of revenue — high concentration risk.",
    "ar": "{name} يمثل {pct}% من الإيرادات — تركز عالٍ.",
    "tr": "{name} gelirin {pct}%'ini oluşturuyor — yüksek konsantrasyon riski."
  },
  "exp_flag_opex_high": {
    "en": "Operating expenses at {pct}% of revenue — review fixed costs.",
    "ar": "مصروفات التشغيل {pct}% من الإيرادات — راجع التكاليف الثابتة.",
    "tr": "Faaliyet giderleri gelirin {pct}%'i — sabit maliyetleri gözden geçirin."
  },
  "exp_impact_high": {
    "en": "High",
    "ar": "عالي",
    "tr": "Yüksek"
  },
  "exp_impact_medium": {
    "en": "Medium",
    "ar": "متوسط",
    "tr": "Orta"
  },
  "exp_impact_low": {
    "en": "Low",
    "ar": "منخفض",
    "tr": "Düşük"
  },
  "exp_dir_up": {
    "en": "Rising",
    "ar": "متصاعد",
    "tr": "Artıyor"
  },
  "exp_dir_down": {
    "en": "Declining",
    "ar": "متراجع",
    "tr": "Düşüyor"
  },
  "exp_dir_stable": {
    "en": "Stable",
    "ar": "مستقر",
    "tr": "Stabil"
  },
  "exp_cat_cogs_goods": {
    "en": "Cost of Goods",
    "ar": "تكلفة البضاعة",
    "tr": "Satılan Mal Maliyeti"
  },
  "exp_cat_labor_direct": {
    "en": "Direct Labor",
    "ar": "العمالة المباشرة",
    "tr": "Doğrudan İşçilik"
  },
  "exp_cat_freight": {
    "en": "Freight & Logistics",
    "ar": "الشحن واللوجستيات",
    "tr": "Nakliye ve Lojistik"
  },
  "exp_cat_salaries": {
    "en": "Salaries & Wages",
    "ar": "الرواتب والأجور",
    "tr": "Maaşlar ve Ücretler"
  },
  "exp_cat_rent": {
    "en": "Rent & Facilities",
    "ar": "الإيجار والمرافق",
    "tr": "Kira ve Tesis"
  },
  "exp_cat_utilities": {
    "en": "Utilities",
    "ar": "الكهرباء والمياه",
    "tr": "Elektrik ve Su"
  },
  "exp_cat_marketing": {
    "en": "Marketing & Advertising",
    "ar": "التسويق والإعلان",
    "tr": "Pazarlama ve Reklam"
  },
  "exp_cat_depreciation": {
    "en": "Depreciation",
    "ar": "الاستهلاك",
    "tr": "Amortisman"
  },
  "exp_cat_maintenance": {
    "en": "Maintenance & Repairs",
    "ar": "الصيانة والإصلاح",
    "tr": "Bakım ve Onarım"
  },
  "exp_cat_insurance": {
    "en": "Insurance & Licenses",
    "ar": "التأمين والتراخيص",
    "tr": "Sigorta ve Lisanslar"
  },
  "exp_cat_admin": {
    "en": "General & Admin",
    "ar": "المصروفات الإدارية",
    "tr": "Genel ve İdari Giderler"
  },
  "exp_cat_tax": {
    "en": "Tax",
    "ar": "الضرائب",
    "tr": "Vergiler"
  },
  "exp_cat_other": {
    "en": "Other Expenses",
    "ar": "مصروفات أخرى",
    "tr": "Diğer Giderler"
  },
  "exp_top_3_title": {
    "en": "Top Cost Drivers",
    "ar": "أكبر محركات التكاليف",
    "tr": "En Büyük Maliyet Kalemleri"
  },
  "exp_of_revenue": {
    "en": "of revenue",
    "ar": "من الإيرادات",
    "tr": "gelirin"
  },
  "exp_of_total": {
    "en": "of total costs",
    "ar": "من إجمالي التكاليف",
    "tr": "toplam maliyetin"
  },
  "exp_mom_up": {
    "en": "↑ Rising",
    "ar": "↑ يرتفع",
    "tr": "↑ Artıyor"
  },
  "exp_mom_down": {
    "en": "↓ Falling",
    "ar": "↓ ينخفض",
    "tr": "↓ Azalıyor"
  },
  "exp_mom_stable": {
    "en": "→ Stable",
    "ar": "→ مستقر",
    "tr": "→ Stabil"
  },
  "exp_cogs_vs_opex": {
    "en": "Cost Structure",
    "ar": "هيكل التكاليف",
    "tr": "Maliyet Yapısı"
  },
  "exp_trend_improving": {
    "en": "Improving",
    "ar": "يتحسن",
    "tr": "İyileşiyor"
  },
  "exp_trend_worsening": {
    "en": "Increasing pressure",
    "ar": "ضغط متزايد",
    "tr": "Baskı artıyor"
  },
  "tab_decisions": {
    "en": "AI Decisions",
    "ar": "قرارات الذكاء الاصطناعي",
    "tr": "AI Kararları"
  },
  "exp_top_rising": {
    "en": "Fastest Growing Cost",
    "ar": "التكلفة الأسرع نمواً",
    "tr": "En Hızlı Büyüyen Maliyet"
  },
  "exp_spike_warning": {
    "en": "Abnormal spike detected",
    "ar": "ارتفاع غير معتاد",
    "tr": "Anormal artış tespit edildi"
  },
  "exp_spike_pct": {
    "en": "jumped",
    "ar": "ارتفع بنسبة",
    "tr": "arttı"
  },
  "exp_vs_rev_label": {
    "en": "% of Revenue",
    "ar": "من الإيرادات",
    "tr": "Gelirin %'si"
  },
  "exp_exec_summary": {
    "en": "Cost Intelligence",
    "ar": "ذكاء التكاليف",
    "tr": "Maliyet Zekası"
  },
  "exp_main_driver": {
    "en": "Main cost driver this period",
    "ar": "المحرك الرئيسي للتكاليف هذه الفترة",
    "tr": "Bu dönemin ana maliyet sürücüsü"
  },
  "exp_no_spike": {
    "en": "No unusual cost movements detected",
    "ar": "لا توجد تحركات غير اعتيادية في التكاليف",
    "tr": "Olağandışı maliyet hareketi tespit edilmedi"
  },
  "exp_account": {
    "en": "Account",
    "ar": "الحساب",
    "tr": "Hesap"
  },
  "exp_amount": {
    "en": "Amount",
    "ar": "المبلغ",
    "tr": "Tutar"
  },
  "exp_breakdown": {
    "en": "Expense Breakdown",
    "ar": "تفصيل المصروفات",
    "tr": "Gider Dağılımı"
  },
  "exp_cogs_label": {
    "en": "COGS",
    "ar": "تكلفة البضاعة",
    "tr": "SMM"
  },
  "exp_opex_label": {
    "en": "OpEx",
    "ar": "مصروفات تشغيل",
    "tr": "Faaliyet Gideri"
  },
  "exp_pct_rev": {
    "en": "% of Rev",
    "ar": "% من الإيراد",
    "tr": "Gelirin %'si"
  },
  "exp_period_label": {
    "en": "Period",
    "ar": "الفترة",
    "tr": "Dönem"
  },
  "exp_structure": {
    "en": "Structure",
    "ar": "هيكل التكاليف",
    "tr": "Yapı"
  },
  "exp_top_drivers": {
    "en": "Top Expense Drivers",
    "ar": "أبرز محركات المصروفات",
    "tr": "Ana Maliyet Kalemleri"
  },
  "exp_total_load": {
    "en": "Total Cost Load",
    "ar": "إجمالي عبء التكاليف",
    "tr": "Toplam Maliyet Yükü"
  },
  "exp_upload_hint": {
    "en": "Upload trial balance to see expense analysis",
    "ar": "ارفع ميزان المراجعة لعرض تحليل المصروفات",
    "tr": "Gider analizi için mizan yükleyin"
  },
  "exp_vs_prior": {
    "en": "vs Prior",
    "ar": "مقارنة بالسابق",
    "tr": "Önceki Dönem"
  },
  "al_ytd_label": {
    "en": "YTD",
    "ar": "منذ بداية العام",
    "tr": "YBD"
  },
  "al_fy_partial_badge": {
    "en": "partial",
    "ar": "جزئي",
    "tr": "kısmi"
  },
  "al_fy_gaps_badge": {
    "en": "gaps",
    "ar": "فجوات",
    "tr": "boşluk"
  },
  "al_kpi_revenue": {
    "en": "Revenue",
    "ar": "الإيرادات",
    "tr": "Gelir"
  },
  "al_kpi_net_profit": {
    "en": "Net Profit",
    "ar": "صافي الربح",
    "tr": "Net Kar"
  },
  "al_kpi_margin": {
    "en": "Margin",
    "ar": "الهامش",
    "tr": "Marj"
  },
  "al_fy_title": {
    "en": "Full Year",
    "ar": "السنة الكاملة",
    "tr": "Tam Yıl"
  },
  "wi_basis_ytd": {
    "en": "YTD",
    "ar": "منذ بداية العام",
    "tr": "YBD"
  },
  "wi_basis_month": {
    "en": "Latest Month",
    "ar": "آخر شهر",
    "tr": "Son Ay"
  },
  "wi_basis_fy": {
    "en": "Full Year",
    "ar": "السنة الكاملة",
    "tr": "Tam Yıl"
  },
  "wi_baseline": {
    "en": "Baseline",
    "ar": "الأساس",
    "tr": "Temel"
  },
  "wi_scenario": {
    "en": "Scenario",
    "ar": "السيناريو",
    "tr": "Senaryo"
  },
  "wi_impact": {
    "en": "Impact",
    "ar": "التأثير",
    "tr": "Etki"
  },
  "wi_revenue": {
    "en": "Revenue",
    "ar": "الإيرادات",
    "tr": "Gelir"
  },
  "wi_net_margin": {
    "en": "Net Margin",
    "ar": "صافي الهامش",
    "tr": "Net Marj"
  },
  "dec_best_action": {
    "en": "Best Action",
    "ar": "أفضل إجراء",
    "tr": "En İyi Eylem"
  },
  "dec_ranking": {
    "en": "Scenario Ranking",
    "ar": "ترتيب السيناريوهات",
    "tr": "Senaryo Sıralaması"
  },
  "dec_score": {
    "en": "Score",
    "ar": "النتيجة",
    "tr": "Puan"
  },
  "dec_np_delta": {
    "en": "Net Profit Δ",
    "ar": "تغيير صافي الربح",
    "tr": "Net Kar Δ"
  },
  "dec_margin_pp": {
    "en": "Margin pp",
    "ar": "نقطة الهامش",
    "tr": "Marj pp"
  },
  "dec_sc_combined": {
    "en": "Accelerate Growth",
    "ar": "تسريع النمو",
    "tr": "Büyümeyi Hızlandır"
  },
  "dec_sc_increase_revenue": {
    "en": "Grow Revenue",
    "ar": "تنمية الإيرادات",
    "tr": "Geliri Artır"
  },
  "dec_sc_reduce_cogs": {
    "en": "Optimize Cost of Goods Sold",
    "ar": "تحسين تكلفة المبيعات",
    "tr": "Satış Maliyetini Optimize Et"
  },
  "dec_sc_reduce_opex": {
    "en": "Reduce Operating Expenses",
    "ar": "خفض المصروفات التشغيلية",
    "tr": "İşletme Giderlerini Azalt"
  },
  "nar_summary": {
    "en": "Executive Summary",
    "ar": "الملخص التنفيذي",
    "tr": "Yönetici Özeti"
  },
  "nar_takeaways": {
    "en": "Key Takeaways",
    "ar": "النقاط الرئيسية",
    "tr": "Temel Çıkarımlar"
  },
  "nar_risks": {
    "en": "Risks & Warnings",
    "ar": "المخاطر والتحذيرات",
    "tr": "Riskler ve Uyarılar"
  },
  "nar_action": {
    "en": "Recommended Action",
    "ar": "الإجراء الموصى به",
    "tr": "Önerilen Eylem"
  },
  "nar_status_excellent": {
    "en": "Excellent",
    "ar": "ممتاز",
    "tr": "Mükemmel"
  },
  "nar_status_good": {
    "en": "Good",
    "ar": "جيد",
    "tr": "İyi"
  },
  "nar_status_warning": {
    "en": "Needs Attention",
    "ar": "يحتاج اهتمام",
    "tr": "Dikkat Gerekiyor"
  },
  "nar_status_critical": {
    "en": "Critical",
    "ar": "حرج",
    "tr": "Kritik"
  },
  "nar_status_neutral": {
    "en": "Neutral",
    "ar": "محايد",
    "tr": "Nötr"
  },
  "exp_excel": {
    "en": "Export Excel",
    "ar": "تصدير Excel",
    "tr": "Excel İndir"
  },
  "exp_json": {
    "en": "Export Report",
    "ar": "تصدير التقرير",
    "tr": "Raporu İndir"
  },
  "exp_loading": {
    "en": "Exporting...",
    "ar": "جاري التصدير...",
    "tr": "İndiriliyor..."
  },
  "exp_sheet_summary": {
    "en": "Summary",
    "ar": "ملخص",
    "tr": "Özet"
  },
  "exp_sheet_annual": {
    "en": "Annual",
    "ar": "سنوي",
    "tr": "Yıllık"
  },
  "exp_sheet_decisions": {
    "en": "Decisions",
    "ar": "القرارات",
    "tr": "Kararlar"
  },
  "exp_sheet_whatif": {
    "en": "What-If",
    "ar": "ماذا لو",
    "tr": "Senaryo"
  },
  "exp_sheet_narrative": {
    "en": "Narrative",
    "ar": "التقرير",
    "tr": "Anlatı"
  },
  "exp_company": {
    "en": "Company",
    "ar": "الشركة",
    "tr": "Şirket"
  },
  "exp_currency": {
    "en": "Currency",
    "ar": "العملة",
    "tr": "Para Birimi"
  },
  "exp_report_basis": {
    "en": "Report Basis",
    "ar": "أساس التقرير",
    "tr": "Rapor Temeli"
  },
  "exp_period": {
    "en": "Period",
    "ar": "الفترة",
    "tr": "Dönem"
  },
  "exp_status": {
    "en": "Status",
    "ar": "الحالة",
    "tr": "Durum"
  },
  "exp_latest_month": {
    "en": "Latest Month",
    "ar": "آخر شهر",
    "tr": "Son Ay"
  },
  "exp_metric": {
    "en": "Metric",
    "ar": "المقياس",
    "tr": "Metrik"
  },
  "exp_change": {
    "en": "Change",
    "ar": "التغيير",
    "tr": "Değişim"
  },
  "exp_rank": {
    "en": "Rank",
    "ar": "الترتيب",
    "tr": "Sıra"
  },
  "exp_scenario": {
    "en": "Scenario",
    "ar": "السيناريو",
    "tr": "Senaryo"
  },
  "exp_revenue_pct": {
    "en": "Revenue %",
    "ar": "الإيرادات %",
    "tr": "Gelir %"
  },
  "exp_cogs_pct": {
    "en": "COGS %",
    "ar": "تكلفة البضاعة %",
    "tr": "SMMM %"
  },
  "exp_opex_pct": {
    "en": "OpEx %",
    "ar": "مصاريف التشغيل %",
    "tr": "OpGider %"
  },
  "exp_operating_expenses": {
    "en": "Operating Expenses",
    "ar": "مصاريف التشغيل",
    "tr": "İşletme Giderleri"
  },
  "exp_gross_margin": {
    "en": "Gross Margin",
    "ar": "هامش الربح الإجمالي",
    "tr": "Brüt Marj"
  },
  "exp_op_margin": {
    "en": "Operating Margin",
    "ar": "هامش التشغيل",
    "tr": "Faaliyet Marjı"
  },
  "exp_op_profit": {
    "en": "Operating Profit",
    "ar": "الربح التشغيلي",
    "tr": "Faaliyet Karı"
  },
  "exp_gross_profit": {
    "en": "Gross Profit",
    "ar": "الربح الإجمالي",
    "tr": "Brüt Kar"
  },
  "exp_annualized": {
    "en": "Partial Year",
    "ar": "سنة جزئية",
    "tr": "Kısmi Yıl"
  },
  "exp_sim_inputs": {
    "en": "Simulation Inputs",
    "ar": "مدخلات المحاكاة",
    "tr": "Simülasyon Girdileri"
  },
  "exp_baseline_vs": {
    "en": "Baseline vs Scenario",
    "ar": "الأساس مقابل السيناريو",
    "tr": "Temel ve Senaryo"
  },
  "exp_best_action": {
    "en": "Best Action",
    "ar": "أفضل إجراء",
    "tr": "En İyi Eylem"
  },
  "exp_full_ranking": {
    "en": "Full Ranking",
    "ar": "الترتيب الكامل",
    "tr": "Tam Sıralama"
  },
  "exp_vcfo_report": {
    "en": "VCFO — Executive Report",
    "ar": "VCFO — التقرير التنفيذي",
    "tr": "VCFO — Yönetici Raporu"
  },
  "exp_no_data": {
    "en": "No simulation data available",
    "ar": "لا تتوفر بيانات محاكاة",
    "tr": "Simülasyon verisi mevcut değil"
  },
  "exp_no_risks": {
    "en": "No risks identified",
    "ar": "لم يتم تحديد مخاطر",
    "tr": "Risk tespit edilmedi"
  }
}


def _t(key: str, lang: str = "en") -> str:
    """Translate a key for the given language. Falls back to English."""
    entry = _TR.get(key)
    if not entry:
        return key
    return entry.get(lang) or entry.get("en") or key


def _align_for_lang(lang: str, h: str = "left") -> Alignment:
    """RTL for Arabic, LTR otherwise."""
    if lang == "ar":
        return Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)
    return Alignment(horizontal=h, vertical="center", wrap_text=True)


# ──────────────────────────────────────────────────────────────────────────────
#  Theme colours
# ──────────────────────────────────────────────────────────────────────────────

DARK_BG   = "0D1829"
ACCENT    = "00D4AA"
HEADER_BG = "0A1F3A"
ALT_ROW   = "0F2440"
RED_CLR   = "FF4D6D"
GREEN_CLR = "10D98A"
AMBER_CLR = "F5A623"
VIOLET    = "7C5CFC"
WHITE     = "E2ECFF"
MUTED     = "374560"


# ──────────────────────────────────────────────────────────────────────────────
#  Style helpers
# ──────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color=WHITE, italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _thin_border() -> Border:
    s = Side(style="thin", color="1A2744")
    return Border(left=s, right=s, top=s, bottom=s)


def _fmtN(v) -> str:
    """Format number with commas, 2 decimal places."""
    if v is None:
        return "—"
    try:
        f = float(v)
        if abs(f) >= 1_000_000:
            return f"{f/1_000_000:,.2f}M"
        if abs(f) >= 1_000:
            return f"{f/1_000:,.1f}K"
        return f"{f:,.0f}"
    except (TypeError, ValueError):
        return str(v)


def _fmtP(v) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v):.2f}%"
    except (TypeError, ValueError):
        return str(v)


def _fmtPP(v) -> str:
    if v is None:
        return "—"
    try:
        f = float(v)
        sign = "+" if f >= 0 else ""
        return f"{sign}{f:.2f} pp"
    except (TypeError, ValueError):
        return str(v)


def _fmtChg(v) -> str:
    if v is None:
        return "—"
    try:
        f = float(v)
        sign = "+" if f >= 0 else ""
        return f"{sign}{f:.1f}%"
    except (TypeError, ValueError):
        return str(v)


# ──────────────────────────────────────────────────────────────────────────────
#  Sheet builders
# ──────────────────────────────────────────────────────────────────────────────

def _style_header_row(ws, row: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(HEADER_BG)
        cell.font = _font(bold=True, size=10, color=ACCENT)
        cell.alignment = _center()
        cell.border = _thin_border()


def _style_title_cell(ws, row: int, col: int, value: str, bg=DARK_BG):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(bg)
    cell.font = _font(bold=True, size=13, color=ACCENT)
    cell.alignment = _left()
    return cell


def _kv_row(ws, row: int, label: str, value, label_color=WHITE, value_color=ACCENT):
    lc = ws.cell(row=row, column=1, value=label)
    lc.fill = _fill(DARK_BG)
    lc.font = _font(bold=True, size=10, color=label_color)
    lc.alignment = _left()
    lc.border = _thin_border()

    vc = ws.cell(row=row, column=2, value=value)
    vc.fill = _fill(ALT_ROW)
    vc.font = _font(size=10, color=value_color)
    vc.alignment = _right()
    vc.border = _thin_border()
    return row + 1


def _blank_row(ws, row: int, fill_color=DARK_BG):
    ws.row_dimensions[row].height = 6
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = _fill(fill_color)
    return row + 1


def _section_title(ws, row: int, title: str, cols=6) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = _fill(HEADER_BG)
    cell.font = _font(bold=True, size=11, color=ACCENT)
    cell.alignment = _left()
    cell.border = _thin_border()
    return row + 1


def _build_summary_sheet(ws, company: dict, basis: str, period_label: str,
                          status: str, annual: dict, currency: str, lang: str = 'en'):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ACCENT[:6]

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    r = 1
    _style_title_cell(ws, r, 1, "VCFO — Executive Report", DARK_BG)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    r = _blank_row(ws, r)

    # Company info
    r = _section_title(ws, r, "Company", cols=4)
    r = _kv_row(ws, r, "Company Name", company.get("name", "—"))
    r = _kv_row(ws, r, "Currency",     currency or "—")
    r = _kv_row(ws, r, "Report Basis", basis)
    r = _kv_row(ws, r, "Period",       period_label)
    r = _kv_row(ws, r, "Status",       status.upper() if status else "—",
                value_color=(GREEN_CLR if status=="excellent" else
                             ACCENT   if status=="good"      else
                             AMBER_CLR if status=="warning"  else
                             RED_CLR  if status=="critical"  else WHITE))
    r = _blank_row(ws, r)

    # Latest month KPIs
    latest = annual.get("latest_month") or {}
    if latest:
        r = _section_title(ws, r, f"{_t('exp_latest_month',lang)} — {latest.get('period','')}", cols=4)
        r = _kv_row(ws, r, "Revenue",           _fmtN(latest.get("revenue")))
        r = _kv_row(ws, r, "COGS",              _fmtN(latest.get("cogs")))
        r = _kv_row(ws, r, "Gross Profit",      _fmtN(latest.get("gross_profit")))
        r = _kv_row(ws, r, "Net Profit",        _fmtN(latest.get("net_profit")))
        r = _kv_row(ws, r, "Gross Margin",      _fmtP(latest.get("gross_margin_pct")))
        r = _kv_row(ws, r, "Net Margin",        _fmtP(latest.get("net_margin_pct")))
        r = _blank_row(ws, r)

    # YTD KPIs
    ytd = annual.get("ytd") or {}
    if ytd:
        r = _section_title(ws, r, f"{_t('al_ytd_label',lang)} {ytd.get('year','')} — {ytd.get('month_count',0)}m", cols=4)
        r = _kv_row(ws, r, "Revenue",     _fmtN(ytd.get("revenue")))
        r = _kv_row(ws, r, "Net Profit",  _fmtN(ytd.get("net_profit")))
        r = _kv_row(ws, r, "Net Margin",  _fmtP(ytd.get("net_margin_pct")))

        ytd_prior = annual.get("ytd_prior") or {}
        comp_ytd  = (annual.get("comparisons") or {}).get("ytd_vs_prior_ytd") or {}
        chg       = (comp_ytd.get("changes") or {})
        if ytd_prior:
            r = _kv_row(ws, r, f"vs {ytd_prior.get('year','')} {_t('al_kpi_revenue',lang)} Δ",
                        _fmtChg(chg.get("revenue")),
                        value_color=GREEN_CLR if (chg.get("revenue") or 0) >= 0 else RED_CLR)
            r = _kv_row(ws, r, f"vs {ytd_prior.get('year','')} {_t('al_kpi_net_profit',lang)} Δ",
                        _fmtChg(chg.get("net_profit")),
                        value_color=GREEN_CLR if (chg.get("net_profit") or 0) >= 0 else RED_CLR)


def _build_annual_sheet(ws, annual: dict, currency: str, lang: str = 'en'):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    for col in ["B","C","D","E","F"]:
        ws.column_dimensions[col].width = 18

    r = 1
    _style_title_cell(ws, r, 1, "Annual & YTD Analysis")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 2

    METRICS = ["revenue","cogs","expenses","gross_profit","operating_profit",
               "net_profit","gross_margin_pct","net_margin_pct","operating_margin_pct"]
    LABELS  = ["Revenue","COGS","OpEx","Gross Profit","Operating Profit",
               "Net Profit",_t("exp_gross_margin", lang),_t("wi_net_margin", lang),"Op Margin %"]
    FMT     = [_fmtN,_fmtN,_fmtN,_fmtN,_fmtN,_fmtN,_fmtP,_fmtP,_fmtP]

    # Build column list: YTD, YTD Prior, then Full Years
    ytd        = annual.get("ytd")       or {}
    ytd_prior  = annual.get("ytd_prior") or {}
    fy_list    = annual.get("full_years") or []

    col_blocks = []
    if ytd:
        col_blocks.append((f"{_t('al_ytd_label',lang)} {ytd.get('year','')} ({ytd.get('month_count',0)}m)", ytd))
    if ytd_prior:
        col_blocks.append((f"{_t('al_ytd_label',lang)} {ytd_prior.get('year','')} ({ytd_prior.get('month_count',0)}m)", ytd_prior))
    for fy in fy_list:
        label = f"{_t('al_fy_title',lang)} {fy['year']}" + ("" if fy.get("complete") else f" ({_t('al_fy_partial_badge',lang)})")
        col_blocks.append((label, fy))

    # Header row
    ws.cell(row=r, column=1, value=_t("exp_metric", lang)).font = _font(bold=True, size=10, color=ACCENT)
    ws.cell(row=r, column=1).fill = _fill(HEADER_BG)
    ws.cell(row=r, column=1).border = _thin_border()
    for ci, (lbl, _) in enumerate(col_blocks, 2):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.fill = _fill(HEADER_BG)
        cell.font = _font(bold=True, size=10, color=ACCENT)
        cell.alignment = _center()
        cell.border = _thin_border()
    r += 1

    # Data rows
    for i, (metric, label, fmt) in enumerate(zip(METRICS, LABELS, FMT)):
        fill_color = DARK_BG if i % 2 == 0 else ALT_ROW
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = _fill(fill_color); lc.font = _font(size=10); lc.border = _thin_border()
        for ci, (_, block) in enumerate(col_blocks, 2):
            val = fmt(block.get(metric))
            vc = ws.cell(row=r, column=ci, value=val)
            vc.fill = _fill(fill_color)
            vc.font = _font(size=10, color=(
                GREEN_CLR if metric in ("gross_profit","net_profit","operating_profit")
                             and (block.get(metric) or 0) > 0
                else RED_CLR if metric in ("gross_profit","net_profit","operating_profit")
                               and (block.get(metric) or 0) < 0
                else WHITE))
            vc.alignment = _right()
            vc.border = _thin_border()
        r += 1


def _build_decisions_sheet(ws, decisions: dict, lang: str = 'en'):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    for col in ["B","C","D","E","F"]:
        ws.column_dimensions[col].width = 16

    r = 1
    _style_title_cell(ws, r, 1, _t("tab_decisions", lang))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 2

    best    = decisions.get("best_scenario") or {}
    ranking = decisions.get("ranking") or []

    if best:
        sc_names = {
            "combined":         _t("dec_sc_combined", lang) + " (+5% / -2% / -2%)",
            "increase_revenue": _t("dec_sc_increase_revenue", lang) + " (+5%)",
            "reduce_cogs":      _t("dec_sc_reduce_cogs", lang) + " (-2%)",
            "reduce_opex":      _t("dec_sc_reduce_opex", lang) + " (-2%)",
        }
        r = _section_title(ws, r, _t("exp_best_action", lang), cols=5)
        r = _kv_row(ws, r, _t("exp_scenario", lang),        sc_names.get(best.get("id",""), best.get("id","—")))
        r = _kv_row(ws, r, _t("dec_score", lang),            f"{best.get('score',0)}/100")
        r = _kv_row(ws, r, _t("dec_np_delta", lang),    _fmtN((best.get("impact") or {}).get("net_profit_delta")),
                    value_color=GREEN_CLR)
        r = _kv_row(ws, r, _t("al_kpi_net_profit", lang) + " %",    _fmtChg((best.get("impact") or {}).get("net_profit_pct_change")),
                    value_color=GREEN_CLR)
        r = _kv_row(ws, r, _t("dec_margin_pp", lang),       _fmtPP((best.get("impact") or {}).get("net_margin_pp")),
                    value_color=GREEN_CLR)
        r = _blank_row(ws, r)

    if ranking:
        r = _section_title(ws, r, _t("exp_full_ranking", lang), cols=5)
        headers = [_t("exp_rank",lang), _t("exp_scenario",lang), _t("dec_score",lang), _t("al_kpi_net_profit",lang)+"%", _t("dec_margin_pp",lang)]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.fill = _fill(HEADER_BG); cell.font = _font(bold=True, size=10, color=ACCENT)
            cell.alignment = _center(); cell.border = _thin_border()
        r += 1
        sc_short = {
            "combined":         _t("dec_sc_combined", lang),
            "increase_revenue": _t("dec_sc_increase_revenue", lang),
            "reduce_cogs":      _t("dec_sc_reduce_cogs", lang),
            "reduce_opex":      _t("dec_sc_reduce_opex", lang),
        }
        for i, sc in enumerate(ranking):
            bg = ALT_ROW if i % 2 else DARK_BG
            vals = [
                sc.get("rank"),
                sc_short.get(sc.get("id",""), sc.get("id","—")),
                f"{sc.get('score',0)}/100",
                _fmtChg(sc.get("np_pct_change")),
                _fmtPP(sc.get("margin_pp")),
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.fill = _fill(bg); cell.font = _font(size=10)
                cell.alignment = _right() if ci > 2 else _left()
                cell.border = _thin_border()
            r += 1


def _build_whatif_sheet(ws, what_if: dict, lang: str = 'en'):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    for col in ["B","C","D"]:
        ws.column_dimensions[col].width = 18

    r = 1
    _style_title_cell(ws, r, 1, _t("exp_sheet_whatif", lang))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    if not what_if or what_if.get("error"):
        ws.cell(row=r, column=1, value=_t("exp_no_data", lang))
        ws.cell(row=r, column=1).font = _font(size=10, color=MUTED, italic=True)
        return

    inputs   = what_if.get("inputs")   or {}
    baseline = what_if.get("baseline") or {}
    scenario = what_if.get("scenario") or {}
    impact   = what_if.get("impact")   or {}

    # Inputs
    r = _section_title(ws, r, _t("exp_sim_inputs", lang), cols=4)
    r = _kv_row(ws, r, _t("exp_revenue_pct", lang),  f"{inputs.get('revenue_pct',0):+.1f}%")
    r = _kv_row(ws, r, _t("exp_cogs_pct", lang),     f"{inputs.get('cogs_pct',0):+.1f}%")
    r = _kv_row(ws, r, _t("exp_opex_pct", lang),     f"{inputs.get('opex_pct',0):+.1f}%")
    r = _blank_row(ws, r)

    # Comparison table
    r = _section_title(ws, r, _t("exp_baseline_vs", lang), cols=4)
    headers = [_t("exp_metric",lang), _t("wi_baseline",lang), _t("wi_scenario",lang), _t("exp_change",lang)]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.fill = _fill(HEADER_BG); cell.font = _font(bold=True, size=10, color=ACCENT)
        cell.alignment = _center(); cell.border = _thin_border()
    r += 1

    rows = [
        (_t("al_kpi_revenue",lang),          "revenue",          _fmtN,  "revenue_pct_change",      _fmtChg),
        (_t("cogs",lang),                    "cogs",             _fmtN,  "cogs_delta",              _fmtN),
        (_t("exp_operating_expenses",lang),  "expenses",         _fmtN,  "expenses_delta",          _fmtN),
        (_t("exp_gross_profit",lang),        "gross_profit",     _fmtN,  "gross_profit_pct_change", _fmtChg),
        (_t("al_kpi_net_profit",lang),       "net_profit",       _fmtN,  "net_profit_pct_change",   _fmtChg),
        (_t("exp_gross_margin",lang),        "gross_margin_pct", _fmtP,  "gross_margin_pp",         _fmtPP),
        (_t("wi_net_margin",lang),           "net_margin_pct",   _fmtP,  "net_margin_pp",           _fmtPP),
    ]
    for i, (label, key, fmt, imp_key, imp_fmt) in enumerate(rows):
        bg = DARK_BG if i % 2 == 0 else ALT_ROW
        cells_data = [
            (label,                 WHITE, _left()),
            (fmt(baseline.get(key)), WHITE, _right()),
            (fmt(scenario.get(key)), ACCENT, _right()),
            (imp_fmt(impact.get(imp_key)), GREEN_CLR if (impact.get(imp_key) or 0) >= 0 else RED_CLR, _right()),
        ]
        for ci, (val, color, align) in enumerate(cells_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = _fill(bg); cell.font = _font(size=10, color=color)
            cell.alignment = align; cell.border = _thin_border()
        r += 1


def _build_narrative_sheet(ws, narrative: dict, lang: str = 'en'):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90

    r = 1
    _style_title_cell(ws, r, 1, _t("nar_summary", lang))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 2

    no_risks = _t("exp_no_risks", lang)
    if lang == "ar":
        missing_summary = "لا توجد أدلة كمية كافية لصياغة ملخص تنفيذي موثوق."
        missing_action = "لا يوجد إجراء موصى به بدرجة ثقة كافية من البيانات الحالية."
    elif lang == "tr":
        missing_summary = "Güvenilir bir yönetici özeti üretmek için yeterli nicel kanıt yok."
        missing_action = "Mevcut verilerden yeterli güven düzeyinde türetilmiş bir önerilen eylem yok."
    else:
        missing_summary = "There is not enough quantitative evidence to produce a reliable executive summary."
        missing_action = "No recommended action can be stated with enough confidence from the current data."
    sections = [
        (_t("nar_summary",lang),   narrative.get("executive_summary") or missing_summary),
        (_t("nar_takeaways",lang), "\n".join(f"✓ {t}" for t in (narrative.get("key_takeaways") or []))),
        (_t("nar_risks",lang),     "\n".join(f"⚠ {t}" for t in (narrative.get("risks") or [])) or no_risks),
        (_t("nar_action",lang),    narrative.get("recommended_action") or missing_action),
    ]

    for label, content in sections:
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = _fill(HEADER_BG); lc.font = _font(bold=True, size=10, color=ACCENT)
        lc.alignment = Alignment(vertical="top", wrap_text=True); lc.border = _thin_border()

        cc = ws.cell(row=r, column=2, value=content)
        cc.fill = _fill(ALT_ROW); cc.font = _font(size=11, color=WHITE)
        cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cc.border = _thin_border()

        lines = max(1, content.count("\n") + 1)
        ws.row_dimensions[r].height = max(30, lines * 16)
        r += 1
        r = _blank_row(ws, r)


# ──────────────────────────────────────────────────────────────────────────────
#  Public API
# ──────────────────────────────────────────────────────────────────────────────

def build_excel(
    company:    dict,
    basis:      str,
    period_label: str,
    status:     str,
    annual:     dict,
    decisions:  dict,
    what_if:    dict,
    narrative:  dict,
    currency:   str = "",
    lang:       str = "en",
) -> bytes:
    """Build a complete Excel workbook and return as bytes."""
    wb = Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = _t("exp_sheet_summary", lang)
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = ACCENT
    _build_summary_sheet(ws1, company, basis, period_label, status, annual, currency, lang)

    # Sheet 2: Annual
    ws2 = wb.create_sheet(_t("exp_sheet_annual", lang))
    _build_annual_sheet(ws2, annual, currency, lang)

    # Sheet 3: Decisions
    ws3 = wb.create_sheet(_t("exp_sheet_decisions", lang))
    _build_decisions_sheet(ws3, decisions, lang)

    # Sheet 4: What-If
    ws4 = wb.create_sheet(_t("exp_sheet_whatif", lang))
    _build_whatif_sheet(ws4, what_if, lang)

    # Sheet 5: Narrative
    ws5 = wb.create_sheet(_t("exp_sheet_narrative", lang))
    _build_narrative_sheet(ws5, narrative, lang)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_report_bundle(
    company:      dict,
    basis:        str,
    period_label: str,
    status:       str,
    annual:       dict,
    decisions:    dict,
    what_if:      dict,
    narrative:    dict,
    warnings:     list,
) -> dict:
    """Build the JSON report bundle."""
    return {
        "company":      company,
        "basis":        basis,
        "basis_period": period_label,
        "status":       status,
        "annual":       annual,
        "decisions":    decisions,
        "what_if":      what_if,
        "narrative":    narrative,
        "warnings":     warnings or [],
    }
